#!/usr/bin/env python3
"""
Game Simulator — reads all data from the RTP spreadsheet and runs both
analytical RTP calculation and Monte Carlo simulation.

The Excel is the single source of truth. Any mismatch is flagged as FAIL.

Usage:
    python3 docs/game-simulator.py              # 10M spins (default)
    python3 docs/game-simulator.py 1000000      # 1M spins (faster)
"""

import sys
import random
from math import comb
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm

XLSX = Path(__file__).parent / "total-bet-model-rtp.xlsx"

# ══════════════════════════════════════════════════════════════
# PART 1: READ EXCEL DATA
# ══════════════════════════════════════════════════════════════

wb = load_workbook(XLSX, data_only=False)
ws = wb.active


def cell(r, c):
    return ws.cell(row=r, column=c).value


# ── Config (row 3) ───────────────────────────────────────────

N_REELS = int(cell(3, 2))     # B3
N_ROWS = int(cell(3, 3))      # C3
STRIP_LEN = int(cell(3, 4))   # D3
N_LINES = int(cell(3, 5))     # E3
COMBOS = STRIP_LEN ** N_REELS

# ── Strips (rows 19-50, cols B-F) ────────────────────────────

NAME_TO_IDX = {"King": 0, "Queen": 1, "Wolf": 2, "J": 3, "K": 4, "Q": 5, "A": 6, "Wild": 7}
IDX_TO_NAME = {v: k for k, v in NAME_TO_IDX.items()}
WILD = 7

strips = []
for reel in range(5):
    col = 2 + reel
    strip = []
    for pos in range(STRIP_LEN):
        row = 19 + pos
        name = cell(row, col)
        assert name in NAME_TO_IDX, f"Unknown symbol '{name}' at row {row}, col {col}"
        strip.append(NAME_TO_IDX[name])
    strips.append(strip)

freq = {}
for idx in range(8):
    freq[idx] = [strip.count(idx) for strip in strips]

# ── Paylines (rows 19-28, cols I-M) ─────────────────────────

paylines = []
for i in range(N_LINES):
    row = 19 + i
    pl = [int(cell(row, 9 + j)) for j in range(5)]
    paylines.append(pl)

# ── Paytable (rows 20-27, cols O-R) ─────────────────────────

SYM_ORDER = ["J", "Q", "K", "A", "Wolf", "Queen", "King", "Wild"]
paytable = {}

for i in range(8):
    row = 20 + i
    sym_name = cell(row, 15)
    assert sym_name in NAME_TO_IDX, f"Unknown symbol '{sym_name}' in paytable row {row}"
    idx = NAME_TO_IDX[sym_name]
    paytable[idx] = (float(cell(row, 16)), float(cell(row, 17)), float(cell(row, 18)))

# ── Wild Pays (rows 46-48, col K) ───────────────────────────

wild_pays = {}
for i, k in enumerate([3, 4, 5]):
    wild_pays[k] = float(cell(46 + i, 11))

# ── Bonus config (rows 53-55) ───────────────────────────────

bonus_pools = {}
bonus_skulls = {}
bonus_prizes = {}
for i in range(3):
    row = 53 + i
    tier = int(cell(row, 8))
    prizes_str = str(cell(row, 10))
    bonus_pools[tier] = int(cell(row, 11))
    bonus_skulls[tier] = int(cell(row, 12))
    bonus_prizes[tier] = [int(x.strip()) for x in prizes_str.split(",")]

# ── Print extracted data ─────────────────────────────────────

print(f"{'='*60}")
print(f"  EXCEL GAME DATA")
print(f"{'='*60}")
print(f"\n  {N_REELS} reels x {N_ROWS} rows | Strip len: {STRIP_LEN} | Paylines: {N_LINES}")

SHORT = {3: "L1", 5: "L2", 4: "L3", 6: "L4", 2: "M1", 1: "M2", 0: "H1", 7: "Wi"}
print(f"\n  Strips:")
for r in range(N_REELS):
    syms = " ".join(SHORT[s] for s in strips[r])
    print(f"    R{r+1}: {syms}")

print(f"\n  Paylines:")
for i, pl in enumerate(paylines):
    print(f"    {i}: {pl}")

print(f"\n  Paytable (x totalBet):")
print(f"    {'':8s} {'x3':>6} {'x4':>6} {'x5':>6}")
for sym in SYM_ORDER:
    idx = NAME_TO_IDX[sym]
    p = paytable[idx]
    print(f"    {sym:8s} {p[0]:>6.2f} {p[1]:>6.2f} {p[2]:>6.2f}")

print(f"\n  Wild pays:  3W={wild_pays[3]}x  4W={wild_pays[4]}x  5W={wild_pays[5]}x")

print(f"\n  Bonus chests:")
for i in range(3):
    print(f"    Tier {i} ({i+3}W): prizes={bonus_prizes[i]}, pool={bonus_pools[i]}x, skulls={bonus_skulls[i]}")


# ══════════════════════════════════════════════════════════════
# PART 2: ANALYTICAL RTP (silent — used for expected values)
# ══════════════════════════════════════════════════════════════

w = [freq[WILD][j] / STRIP_LEN for j in range(5)]

total_line_rtp = 0.0
for idx in range(8):
    is_wild = idx == WILD
    p = [(freq[idx][j] + freq[WILD][j]) / STRIP_LEN if not is_wild else freq[WILD][j] / STRIP_LEN for j in range(5)]
    q = [1 - p[j] for j in range(5)]

    for count in [3, 4, 5]:
        if is_wild and count < 5:
            continue
        if is_wild:
            prob = 1
            for j in range(5):
                prob *= w[j]
        else:
            pp = 1
            wp = 1
            for j in range(count):
                pp *= p[j]
                wp *= w[j]
            prob = (pp - wp) * q[count] if count < 5 else pp - wp

        total_line_rtp += prob * N_LINES * paytable[idx][count - 3]

p_wild_grid = freq[WILD][0] * N_ROWS / STRIP_LEN

total_wild_pay_rtp = 0.0
for k in [3, 4, 5]:
    pk = comb(5, k) * p_wild_grid**k * (1 - p_wild_grid)**(5 - k)
    total_wild_pay_rtp += pk * wild_pays[k]

total_bonus_rtp = 0.0
bonus_evs = {}
for i, k in enumerate([3, 4, 5]):
    pk = comb(5, k) * p_wild_grid**k * (1 - p_wild_grid)**(5 - k)
    ev = bonus_pools[i] / (bonus_skulls[i] + 1)
    bonus_evs[i] = ev
    total_bonus_rtp += pk * ev

analytical_rtp = total_line_rtp + total_wild_pay_rtp + total_bonus_rtp


# ══════════════════════════════════════════════════════════════
# PART 3: MONTE CARLO SIMULATION
# ══════════════════════════════════════════════════════════════

NUM_SPINS = int(sys.argv[1]) if len(sys.argv) > 1 else 10_000_000

print(f"\n{'='*60}")
print(f"  MONTE CARLO SIMULATION  ({NUM_SPINS:,} spins)")
print(f"{'='*60}\n")

rng = random.Random(42)

total_wagered = 0.0
mc_line_payout = 0.0
mc_wild_pay_payout = 0.0
mc_bonus_payout = 0.0
mc_winning_spins = 0
mc_bonus_triggers = 0

TOTAL_BET = 1  # normalised

with tqdm(total=NUM_SPINS, desc="Spinning", ncols=100) as pbar:
    for spin in range(NUM_SPINS):
        # Random reel stops -> build 5x3 grid
        grid = []
        wild_count = 0
        for r in range(N_REELS):
            stop = rng.randint(0, STRIP_LEN - 1)
            col = []
            has_wild = False
            for row in range(N_ROWS):
                sym = strips[r][(stop + row) % STRIP_LEN]
                col.append(sym)
                if sym == WILD:
                    has_wild = True
            grid.append(col)
            if has_wild:
                wild_count += 1

        total_wagered += TOTAL_BET

        # Evaluate paylines
        line_win = 0.0
        for pl in paylines:
            syms = [grid[r][pl[r]] for r in range(N_REELS)]

            base = WILD
            for s in syms:
                if s != WILD:
                    base = s
                    break

            count = 0
            for s in syms:
                if s == base or s == WILD:
                    count += 1
                else:
                    break

            if count >= 3:
                pays = paytable.get(base)
                if pays:
                    line_win += pays[count - 3] * TOTAL_BET

        mc_line_payout += line_win

        # Wild pay
        wp_val = wild_pays.get(wild_count, 0) * TOTAL_BET
        mc_wild_pay_payout += wp_val

        has_win = line_win > 0 or wp_val > 0

        # Bonus simulation
        if wild_count >= 3:
            mc_bonus_triggers += 1
            tier = min(wild_count - 3, 2)
            prizes = list(bonus_prizes[tier])
            skulls_count = bonus_skulls[tier]

            chests = [p * TOTAL_BET for p in prizes] + [None] * skulls_count
            rng.shuffle(chests)

            bonus_win = 0.0
            for chest in chests:
                if chest is None:
                    break
                bonus_win += chest

            mc_bonus_payout += bonus_win
            if bonus_win > 0:
                has_win = True

        if has_win:
            mc_winning_spins += 1

        pbar.update(1)

mc_total_payout = mc_line_payout + mc_wild_pay_payout + mc_bonus_payout
mc_rtp = mc_total_payout / total_wagered
mc_line_rtp = mc_line_payout / total_wagered
mc_wp_rtp = mc_wild_pay_payout / total_wagered
mc_bonus_rtp = mc_bonus_payout / total_wagered
mc_hit_rate = mc_winning_spins / NUM_SPINS
mc_bonus_freq = mc_bonus_triggers / NUM_SPINS


# ══════════════════════════════════════════════════════════════
# PART 4: VERIFICATION
# ══════════════════════════════════════════════════════════════

print(f"{'='*60}")
print(f"  VERIFICATION")
print(f"{'='*60}")

# EV checks
expected_evs = [83, 171, 383]
all_ev_ok = True
print(f"\n  EV totals:")
for i, k in enumerate([3, 4, 5]):
    ev_bonus = bonus_evs[i]
    ev_total = ev_bonus + wild_pays[k]
    expected = expected_evs[i]
    ok = abs(ev_total - expected) < 0.01
    status = "PASS" if ok else "FAIL"
    if not ok:
        all_ev_ok = False
    print(f"    {k}W: {wild_pays[k]} + {ev_bonus:.0f} = {ev_total:.0f}  (expected {expected})  {status}")

# Analytical checks
print(f"\n  Analytical RTP:")
analytical_checks = [
    ("Total RTP",          analytical_rtp,                              0.9614),
    ("Line RTP",           total_line_rtp,                              0.3371),
    ("WildPay + Bonus",    total_wild_pay_rtp + total_bonus_rtp,        0.6243),
]

all_pass = all_ev_ok
for label, actual, expected in analytical_checks:
    diff = actual - expected
    ok = abs(diff) < 0.002
    status = "PASS" if ok else "FAIL"
    if not ok:
        all_pass = False
    print(f"    {label:20s}  {actual*100:7.4f}%  (expected ~{expected*100:.2f}%)  err {diff*100:+.4f}%  {status}")

# Data integrity
strip_ok = all(len(s) == STRIP_LEN for s in strips)
freq_ok = all(sum(freq[idx][r] for idx in range(8)) == STRIP_LEN for r in range(5))
if not strip_ok or not freq_ok:
    all_pass = False
print(f"    {'Strip integrity':20s}  {'PASS' if strip_ok and freq_ok else 'FAIL'}")

# Monte Carlo vs analytical
MC_TOLERANCE = 0.01  # 1.0%
print(f"\n  Monte Carlo vs analytical:")

mc_comparisons = [
    ("Total RTP",     mc_rtp,       analytical_rtp),
    ("Line RTP",      mc_line_rtp,  total_line_rtp),
    ("Wild Pay RTP",  mc_wp_rtp,    total_wild_pay_rtp),
    ("Bonus RTP",     mc_bonus_rtp, total_bonus_rtp),
]

for label, mc_val, expected_val in mc_comparisons:
    diff = mc_val - expected_val
    ok = abs(diff) < MC_TOLERANCE
    status = "PASS" if ok else "FAIL"
    if not ok:
        all_pass = False
    print(f"    {label:20s}  {mc_val*100:7.4f}%  (expected {expected_val*100:.4f}%)  err {diff*100:+.4f}%  {status}")

print(f"    {'Hit rate':20s}  {mc_hit_rate*100:.2f}%")
print(f"    {'Bonus frequency':20s}  {mc_bonus_freq*100:.3f}%  ({mc_bonus_triggers:,} triggers)")

print(f"\n{'='*60}")
if all_pass:
    print("  ALL CHECKS PASSED")
else:
    print("  SOME CHECKS FAILED")
print(f"{'='*60}")

sys.exit(0 if all_pass else 1)
