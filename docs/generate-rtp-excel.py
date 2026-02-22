"""
Generates Total Bet Model RTP spreadsheet — single sheet, flat layout.
Matches the visual pattern of slotMaths(1).xlsx.
Run: python3 docs/generate-rtp-excel.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Full 1"

# ── Styles ────────────────────────────────────────────────────

BLD = Font(bold=True)
BLD12 = Font(bold=True, size=12)
SZ8 = Font(size=8, italic=True, color="888888")
TB = Border(*(Side(style="thin"),) * 4)
WILD_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HDR_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
PCT = "0.0000%"
PCT2 = "0.00%"
M2 = "0.00"
M6 = "0.000000"
INT = "#,##0"


def b(ws, r, c, val=None, fmt=None, bold=False, fill=None):
    """Set cell value with optional formatting."""
    cl = ws.cell(row=r, column=c, value=val)
    if bold:
        cl.font = BLD
    if fmt:
        cl.number_format = fmt
    if fill:
        cl.fill = fill
    return cl


def hdr_row(ws, r, labels, start_col=1, fill=None):
    """Write a row of bold headers."""
    f = fill or HDR_FILL
    for i, lbl in enumerate(labels):
        cl = ws.cell(row=r, column=start_col + i, value=lbl)
        cl.font = BLD
        cl.fill = f


# ── Data ──────────────────────────────────────────────────────

IDX_TO_NAME = {0: "King", 1: "Queen", 2: "Wolf", 3: "J", 4: "K", 5: "Q", 6: "A", 7: "Wild"}

STRIPS_RAW = [
    [0,3,1,4,2,5,0,6,3,1,5,2,4,6,0,3,1,7,5,2],
    [1,4,0,5,3,6,2,0,4,1,3,5,6,2,0,4,5,1,3,7],
    [2,5,3,0,6,1,4,2,5,3,0,6,1,4,7,2,5,3,0,6],
    [3,0,4,1,5,2,6,3,0,4,1,5,2,6,3,0,5,4,1,7],
    [4,1,5,2,0,3,6,4,1,5,2,0,3,6,4,7,1,5,2,0],
]

# Convert strip indices to names
STRIPS = [[IDX_TO_NAME[idx] for idx in reel] for reel in STRIPS_RAW]

PAYLINES = [
    [1,1,1,1,1],[0,0,0,0,0],[2,2,2,2,2],[0,1,2,1,0],[2,1,0,1,2],
    [0,0,1,0,0],[2,2,1,2,2],[1,2,2,2,1],[1,0,0,0,1],[0,1,1,1,0],
    [2,1,1,1,2],[1,0,1,0,1],[1,2,1,2,1],[0,1,0,1,0],[2,1,2,1,2],
    [0,0,1,2,2],[2,2,1,0,0],[1,0,0,1,2],[1,2,2,1,0],[0,1,2,2,1],
]

# Symbol order: low → high (matching slotMaths reference)
SYM_ORDER = ["J", "Q", "K", "A", "Wolf", "Queen", "King", "Wild"]

# Old paytable (x betPerLine)
OLD_PAY = {
    "J":     (2, 4, 14),
    "Q":     (2, 5, 17),
    "K":     (3, 7, 25),
    "A":     (4, 9, 34),
    "Wolf":  (7, 17, 51),
    "Queen": (8, 25, 85),
    "King":  (9, 33, 168),
    "Wild":  (10, 34, 250),
}

WILD_PAYS = {3: 2, 4: 5, 5: 20}
BONUS_PRIZES = {0: [8, 18, 28], 1: [8, 15, 20, 27], 2: [8, 10, 15, 20, 27]}
BONUS_SKULLS = {0: 2, 1: 1, 2: 0}
BUY_PRICES = {0: 21, 1: 42, 2: 105}
OLD_EVS = {0: 20, 1: 40, 2: 100}

STRIP_LEN = 20
N_REELS = 5
N_ROWS = 3
N_LINES = 20


# ── Exact hit rate (brute-force from strip/payline data) ─────

def compute_hit_rate():
    """Exact hit rate: enumerate all stop combos via optimized decomposition.

    Key insight: a line win depends only on the first 3 reels (3+ consecutive
    from left). So we precompute any_line_win[s0][s1][s2] for 20^3 = 8K combos,
    then combine with wild visibility on reels 3-4 analytically. Total work: O(20^3).
    """
    W = 7  # Wild symbol index

    # Precompute reel columns (visible symbols per stop position)
    reel_cols = []
    for r in range(N_REELS):
        cols = []
        for s in range(STRIP_LEN):
            cols.append(tuple(STRIPS_RAW[r][(s + row) % STRIP_LEN] for row in range(N_ROWS)))
        reel_cols.append(cols)

    # Precompute wild visibility per reel/stop (1 if wild in visible area, else 0)
    wild_visible = []
    for r in range(N_REELS):
        wv = [0] * STRIP_LEN
        for s in range(STRIP_LEN):
            for row in range(N_ROWS):
                if STRIPS_RAW[r][(s + row) % STRIP_LEN] == W:
                    wv[s] = 1
                    break
        wild_visible.append(wv)

    # Precompute any_line_win[s0][s1][s2] — True if any payline has 3+ match from left
    any_line_win = [[[False] * STRIP_LEN for _ in range(STRIP_LEN)] for _ in range(STRIP_LEN)]
    for s0 in range(STRIP_LEN):
        c0 = reel_cols[0][s0]
        for s1 in range(STRIP_LEN):
            c1 = reel_cols[1][s1]
            for s2 in range(STRIP_LEN):
                c2 = reel_cols[2][s2]
                for pl in PAYLINES:
                    a, bv, cv = c0[pl[0]], c1[pl[1]], c2[pl[2]]
                    if a != W:
                        base = a
                    elif bv != W:
                        base = bv
                    elif cv != W:
                        base = cv
                    else:
                        any_line_win[s0][s1][s2] = True
                        break
                    if (a == base or a == W) and (bv == base or bv == W) and (cv == base or cv == W):
                        any_line_win[s0][s1][s2] = True
                        break

    # Wild combos on reels 3-4 (analytical — avoids iterating s3,s4)
    n_w3 = sum(wild_visible[3])
    n_w4 = sum(wild_visible[4])
    sl2 = STRIP_LEN ** 2
    wild_34_any = sl2 - (STRIP_LEN - n_w3) * (STRIP_LEN - n_w4)  # at least 1 wild
    wild_34_both = n_w3 * n_w4  # both wilds

    # Count winning combos across all 20^5 stops
    winning = 0
    for s0 in range(STRIP_LEN):
        wv0 = wild_visible[0][s0]
        for s1 in range(STRIP_LEN):
            wv01 = wv0 + wild_visible[1][s1]
            for s2 in range(STRIP_LEN):
                wv012 = wv01 + wild_visible[2][s2]
                if any_line_win[s0][s1][s2] or wv012 >= 3:
                    winning += sl2
                elif wv012 == 2:
                    winning += wild_34_any
                elif wv012 == 1:
                    winning += wild_34_both

    return winning / (STRIP_LEN ** N_REELS)


print("Computing exact hit rate...")
EXACT_HIT_RATE = compute_hit_rate()
print(f"Hit Rate: {EXACT_HIT_RATE * 100:.2f}%")


# ── Column widths ─────────────────────────────────────────────

col_widths = {'A': 4, 'B': 10, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10,
              'H': 5, 'I': 10, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12,
              'O': 14, 'P': 14, 'Q': 14, 'R': 14, 'S': 10, 'T': 10, 'U': 10}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════
# ROW 2-3: CONFIG
# ══════════════════════════════════════════════════════════════

b(ws, 2, 2, "Columns", bold=True)
b(ws, 2, 3, "Rows", bold=True)
b(ws, 2, 4, "Strip Len", bold=True)
b(ws, 2, 5, "Paylines", bold=True)
b(ws, 2, 6, "Total Combos", bold=True)

b(ws, 3, 2, N_REELS)
b(ws, 3, 3, N_ROWS)
b(ws, 3, 4, STRIP_LEN)
b(ws, 3, 5, N_LINES)
b(ws, 3, 6, "=D3^B3", INT)  # F3 = total combos


# ══════════════════════════════════════════════════════════════
# ROW 5-15: FREQUENCIES + PROBS + EV
# ══════════════════════════════════════════════════════════════

# --- Frequency headers (cols B-G) ---
freq_labels = ["Symbol", "Reel 1", "Reel 2", "Reel 3", "Reel 4", "Reel 5"]
hdr_row(ws, 5, freq_labels, 2)

# Strip data range: rows 19-38, cols B-F (defined below, referenced here)
STRIP_R0 = 19  # first strip data row
STRIP_R1 = STRIP_R0 + STRIP_LEN - 1  # last strip data row

# Frequency data (rows 6-13)
FREQ_R0 = 6
for i, sym in enumerate(SYM_ORDER):
    r = FREQ_R0 + i
    b(ws, r, 2, sym, bold=True)
    for j in range(5):
        scol = get_column_letter(2 + j)  # B=reel1 strip col ... F=reel5
        b(ws, r, 3 + j, f'=COUNTIF({scol}${STRIP_R0}:{scol}${STRIP_R1},$B{r})', INT)

# Total row
FREQ_TOTAL_R = FREQ_R0 + len(SYM_ORDER)  # row 14
b(ws, FREQ_TOTAL_R, 2, "Total", bold=True)
for j in range(5):
    col = 3 + j
    b(ws, FREQ_TOTAL_R, col, f"=SUM({get_column_letter(col)}{FREQ_R0}:{get_column_letter(col)}{FREQ_TOTAL_R-1})", INT)

# Wild row reference (Wild is last in SYM_ORDER = row 13)
WILD_SYM_R = FREQ_R0 + SYM_ORDER.index("Wild")  # row 13

# --- Prob per reel headers (cols I-N) ---
prob_labels = ["Symbol", "Prob R1", "Prob R2", "Prob R3", "Prob R4", "Prob R5"]
hdr_row(ws, 5, prob_labels, 9)

# Prob 3/4/5 + EV headers
hdr_row(ws, 5, ["Prob 3", "Prob 4", "Prob 5", "EV"], 15)

# Prob data (rows 6-13)
# p_j = (freq_sym + freq_wild) / total for non-wild
# p_j = freq_wild / total for Wild
# Freq cols: C=R1, D=R2, E=R3, F=R4, G=R5
# Total cols: same row 14
# Prob cols: J=R1, K=R2, L=R3, M=R4, N=R5
PROB_COLS = ['J', 'K', 'L', 'M', 'N']
FREQ_COLS = ['C', 'D', 'E', 'F', 'G']

for i, sym in enumerate(SYM_ORDER):
    r = FREQ_R0 + i
    b(ws, r, 9, sym, bold=True)  # col I

    is_wild = (sym == "Wild")
    for j in range(5):
        fc = FREQ_COLS[j]  # freq column for this reel
        if is_wild:
            # Wild: p = freq_wild / total
            formula = f"={fc}{r}/{fc}${FREQ_TOTAL_R}"
        else:
            # Non-wild: p = (freq_sym + freq_wild) / total
            formula = f"=({fc}{r}+{fc}${WILD_SYM_R})/{fc}${FREQ_TOTAL_R}"
        b(ws, r, 10 + j, formula, M6)  # cols J-N

    # Prob 3, 4, 5 (cols O, P, Q)
    if is_wild:
        # Wild: only ×5 is possible (all 5 must be wild)
        b(ws, r, 15, 0, M6)  # Prob3 = 0
        b(ws, r, 16, 0, M6)  # Prob4 = 0
        # Prob5 = product of w_j (just wild prob per reel, not p_j)
        # w_j = freq_wild / total = same as p_j for Wild row
        w_terms = "*".join(f"{PROB_COLS[j]}{r}" for j in range(5))
        b(ws, r, 17, f"={w_terms}", M6)
    else:
        # Prob3 = (p0*p1*p2 - w0*w1*w2) * q3
        # where w_j = Wild's p_j (row WILD_SYM_R), q3 = 1 - p3
        p3 = "*".join(f"{PROB_COLS[j]}{r}" for j in range(3))
        w3 = "*".join(f"{PROB_COLS[j]}{WILD_SYM_R}" for j in range(3))
        b(ws, r, 15, f"=({p3}-{w3})*(1-{PROB_COLS[3]}{r})", M6)

        # Prob4 = (p0*p1*p2*p3 - w0*w1*w2*w3) * q4
        p4 = "*".join(f"{PROB_COLS[j]}{r}" for j in range(4))
        w4 = "*".join(f"{PROB_COLS[j]}{WILD_SYM_R}" for j in range(4))
        b(ws, r, 16, f"=({p4}-{w4})*(1-{PROB_COLS[4]}{r})", M6)

        # Prob5 = p0*p1*p2*p3*p4 - w0*w1*w2*w3*w4
        p5 = "*".join(f"{PROB_COLS[j]}{r}" for j in range(5))
        w5 = "*".join(f"{PROB_COLS[j]}{WILD_SYM_R}" for j in range(5))
        b(ws, r, 17, f"={p5}-{w5}", M6)

    # EV = pay3*prob3 + pay4*prob4 + pay5*prob5
    # Pay values are in the paytable area (cols P-R, starting at PAY_R0)
    # We'll fill paytable later; for now reference it
    PAY_R0_OFFSET = 20  # paytable data starts at row 20 (defined below)
    pay_r = PAY_R0_OFFSET + i
    b(ws, r, 18, f"=P{pay_r}*O{r}+Q{pay_r}*P{r}+R{pay_r}*Q{r}", M6)


# ── Row 15: Wild visible in grid (for scatter/bonus trigger) ──

b(ws, 15, 9, "Wild visible/reel", bold=True)
for j in range(5):
    fc = FREQ_COLS[j]
    # P(wild visible on reel) = freq_wild * visible_rows / strip_length
    b(ws, 15, 10 + j, f"={fc}{WILD_SYM_R}*$C$3/{fc}${FREQ_TOTAL_R}", M6)

# P(3/4/5 wilds in grid) using binomial with p from J15
b(ws, 15, 15, "P(3W grid)", bold=True)
b(ws, 15, 16, "P(4W grid)", bold=True)
b(ws, 15, 17, "P(5W grid)", bold=True)

# Use p = J15 (same for all reels)
for i, k in enumerate([3, 4, 5]):
    b(ws, 16, 15 + i, f"=COMBIN($B$3,{k})*$J$15^{k}*(1-$J$15)^($B$3-{k})", M6)


# ══════════════════════════════════════════════════════════════
# ROW 18-37: STRIPS + PAYLINES + PAYTABLE
# ══════════════════════════════════════════════════════════════

# --- Strip headers (row 18, data 19-38) ---
hdr_row(ws, 18, ["Strip 1", "Strip 2", "Strip 3", "Strip 4", "Strip 5"], 2)

for pos in range(STRIP_LEN):
    r = STRIP_R0 + pos
    for j in range(5):
        cl = b(ws, r, 2 + j, STRIPS[j][pos])
        if STRIPS[j][pos] == "Wild":
            cl.fill = WILD_FILL
            cl.font = BLD

# --- Payline headers (row 18, cols H-M, data 19-38) ---
hdr_row(ws, 18, ["Line ID", "Reel 1", "Reel 2", "Reel 3", "Reel 4", "Reel 5"], 8)

for i, pl in enumerate(PAYLINES):
    r = STRIP_R0 + i
    b(ws, r, 8, i)  # Line ID (0-19)
    for j in range(5):
        b(ws, r, 9 + j, pl[j])

# --- Paytable (row 19, cols O-R) ---
hdr_row(ws, 19, ["Symbol", "Pay 3", "Pay 4", "Pay 5"], 15)

PAY_R0 = 20  # first paytable data row
for i, sym in enumerate(SYM_ORDER):
    r = PAY_R0 + i
    x3, x4, x5 = OLD_PAY[sym]

    b(ws, r, 15, sym, bold=True)
    # New = Old / Paylines
    b(ws, r, 16, x3 / N_LINES, M2)
    b(ws, r, 17, x4 / N_LINES, M2)
    b(ws, r, 18, x5 / N_LINES, M2)


# ══════════════════════════════════════════════════════════════
# ROW 30-35: RTP SUMMARY
# ══════════════════════════════════════════════════════════════

hdr_row(ws, 30, ["", "RTP", "RTP %"], 15)

LAST_SYM_R = FREQ_R0 + len(SYM_ORDER) - 1  # row 13

# Base RTP = SUM(EV per symbol) × paylines
b(ws, 31, 15, "Base", bold=True)
b(ws, 31, 16, f"=SUM(R{FREQ_R0}:R{LAST_SYM_R})*$E$3", PCT)
b(ws, 31, 17, f"=P31*100", M2)

# Wild Pay RTP (references wild pay section below)
b(ws, 32, 15, "Wild Pay", bold=True)
b(ws, 32, 16, f"=L44", PCT)  # total wild pay RTP from wild pay section
b(ws, 32, 17, f"=P32*100", M2)

# Bonus RTP
b(ws, 33, 15, "Bonus", bold=True)
b(ws, 33, 16, f"=O50", PCT)  # total bonus RTP from bonus section
b(ws, 33, 17, f"=P33*100", M2)

# Total
b(ws, 34, 15, "Total", bold=True)
b(ws, 34, 16, f"=P31+P32+P33", PCT)
b(ws, 34, 17, f"=P34*100", M2)
for c in range(15, 18):
    ws.cell(row=34, column=c).fill = TOT_FILL
    ws.cell(row=34, column=c).font = BLD12

# Hit Rate (exact — brute-force from strip/payline data)
b(ws, 36, 15, "Hit Rate", bold=True)
b(ws, 36, 16, EXACT_HIT_RATE, PCT)
b(ws, 36, 17, f"=P36*100", M2)
b(ws, 37, 15, "(exact)", bold=False)
ws.cell(row=37, column=15).font = SZ8


# ══════════════════════════════════════════════════════════════
# ROW 40-44: WILD PAY (Wild pays for 3+ in grid)
# ══════════════════════════════════════════════════════════════

b(ws, 40, 8, "WILD PAY (Wild en grid)", bold=True)
b(ws, 40, 8).font = BLD12

hdr_row(ws, 41, ["Wilds", "P(k wilds)", "Combinaciones", "Wild Pay (xTBet)", "Wild Pay RTP"], 8)

SC_R0 = 42  # rows 42, 43, 44
for i, k in enumerate([3, 4, 5]):
    r = SC_R0 + i
    b(ws, r, 8, k)
    # P(k wilds) — reference from row 16
    b(ws, r, 9, f"=O{16 + (k-3)}", M6)  # O16=P(3W), P16=P(4W), Q16=P(5W)
    # Wait, O16/P16/Q16 are in cols 15,16,17. Let me fix references.
    # Actually row 16 cols O,P,Q hold P(3W), P(4W), P(5W)

# Fix: P(3W) is at row 16 col 15 (O16), P(4W) at col 16 (P16), P(5W) at col 17 (Q16)
for i, k in enumerate([3, 4, 5]):
    r = SC_R0 + i
    ws.cell(row=r, column=8, value=k)
    pcol = get_column_letter(15 + i)  # O, P, Q
    b(ws, r, 9, f"={pcol}16", M6)
    b(ws, r, 10, f"=I{r}*$F$3", M2)  # occurrences = P × totalCombos
    b(ws, r, 11, WILD_PAYS[k], M2)  # wild pay (design input)
    b(ws, r, 12, f"=I{r}*K{r}", PCT)  # wild pay RTP = P × pay

# Total wild pay
b(ws, SC_R0 + 3, 8, "Total Wild Pay", bold=True)
b(ws, SC_R0 + 3, 10, f"=SUM(J{SC_R0}:J{SC_R0+2})", M2)
b(ws, SC_R0 + 3, 12, f"=SUM(L{SC_R0}:L{SC_R0+2})", PCT)
# This is L44 — referenced by RTP summary O32

# Fix the O32 reference — wild pay total is at L{SC_R0+3} = L45
# Let me recalculate: SC_R0=42, total row = 42+3 = 45
# But I wrote L44 in O32 above. Let me check: 42,43,44 = data, 45 = total. Fix:

ws.cell(row=32, column=16).value = f"=L{SC_R0+3}"  # P32 = total wild pay RTP


# ══════════════════════════════════════════════════════════════
# ROW 47-51: BONUS (Chest pick-me)
# ══════════════════════════════════════════════════════════════

b(ws, 47, 8, "BONUS COFRES", bold=True)
b(ws, 47, 8).font = BLD12

hdr_row(ws, 48, ["Tier", "Wilds", "Prizes (xTBet)", "Pool", "Skulls",
                  "E[picks]", "EV bonus (xTBet)", "Bonus RTP"], 8)
ws.column_dimensions['J'].width = 20

BN_R0 = 49  # rows 49, 50, 51
for i in range(3):
    r = BN_R0 + i
    wilds = 3 + i
    prizes = BONUS_PRIZES[i]
    skulls = BONUS_SKULLS[i]

    b(ws, r, 8, i)               # tier
    b(ws, r, 9, wilds)           # wilds
    b(ws, r, 10, ", ".join(str(p) for p in prizes))  # prizes text
    b(ws, r, 11, sum(prizes))    # pool
    b(ws, r, 12, skulls)         # skulls
    # E[picks] = prizes_count / (skulls + 1)
    # prizes_count: I could count from text but simpler to use (5-skulls)
    b(ws, r, 13, f"=(5-L{r})/(L{r}+1)", M2)
    # EV = pool / (skulls + 1)
    b(ws, r, 14, f"=K{r}/(L{r}+1)", M2)
    # Bonus RTP = P(k wilds) × EV_bonus
    # P(k wilds) is at I{SC_R0+i} in scatter section
    sc_p_ref = f"I{SC_R0 + i}"
    b(ws, r, 15, f"={sc_p_ref}*N{r}", PCT)

# Total bonus RTP
BN_TOT_R = BN_R0 + 3  # row 52
b(ws, BN_TOT_R, 8, "Total Bonus", bold=True)
b(ws, BN_TOT_R, 15, f"=SUM(O{BN_R0}:O{BN_R0+2})", PCT)

# Fix P33 reference to point to correct bonus total
ws.cell(row=33, column=16).value = f"=O{BN_TOT_R}"


# ══════════════════════════════════════════════════════════════
# ROW 47-51 (right side): WILD PAY + BONUS EV VERIFICATION
# ══════════════════════════════════════════════════════════════

hdr_row(ws, 48, ["Wild Pay", "Bonus EV", "Total EV", "Old EV", "Match?"], 17)

for i in range(3):
    r = BN_R0 + i
    sc_pay_ref = f"K{SC_R0 + i}"
    bn_ev_ref = f"N{r}"
    b(ws, r, 17, f"={sc_pay_ref}", M2)           # wild pay
    b(ws, r, 18, f"={bn_ev_ref}", M2)             # bonus EV
    b(ws, r, 19, f"=Q{r}+R{r}", M2)               # total
    b(ws, r, 20, OLD_EVS[i], M2)                   # old EV
    b(ws, r, 21, f'=IF(ABS(S{r}-T{r})<0.01,"OK","FAIL")')


# ══════════════════════════════════════════════════════════════
# ROW 54-58: BUY BONUS
# ══════════════════════════════════════════════════════════════

b(ws, 54, 8, "BUY BONUS", bold=True)
b(ws, 54, 8).font = BLD12

hdr_row(ws, 55, ["Tier", "Simula", "EV total (xTBet)", "Precio (xTBet)", "RTP Buy"], 8)

BB_R0 = 56
for i in range(3):
    r = BB_R0 + i
    wilds = 3 + i
    b(ws, r, 8, f"Tier {i+1}")
    b(ws, r, 9, f"{wilds} Wilds")
    # EV total = scatter + bonus (from verification row)
    b(ws, r, 10, f"=S{BN_R0+i}", M2)
    b(ws, r, 11, BUY_PRICES[i], M2)
    b(ws, r, 12, f"=J{r}/K{r}", PCT)


# ══════════════════════════════════════════════════════════════
# ROW 60-66: VERIFICATION BLOCK
# ══════════════════════════════════════════════════════════════

b(ws, 60, 8, "VERIFICACIONES", bold=True)
b(ws, 60, 8).font = BLD12

hdr_row(ws, 61, ["Check", "Resultado"], 8)

checks = [
    ("RTP total = 96.12%?", f'=IF(ABS(P34-0.961232)<0.0001,"PASS","FAIL")'),
    ("WildPay+Bonus = 58.13%?", f'=IF(ABS(P32+P33-0.581344)<0.0001,"PASS","FAIL")'),
    ("Line RTP = 37.99%?", f'=IF(ABS(P31-0.379888)<0.0001,"PASS","FAIL")'),
    ("EV 3W = 20?", f'=IF(ABS(S{BN_R0}-T{BN_R0})<0.01,"PASS","FAIL")'),
    ("EV 4W = 40?", f'=IF(ABS(S{BN_R0+1}-T{BN_R0+1})<0.01,"PASS","FAIL")'),
    ("EV 5W = 100?", f'=IF(ABS(S{BN_R0+2}-T{BN_R0+2})<0.01,"PASS","FAIL")'),
]

for i, (name, formula) in enumerate(checks):
    r = 62 + i
    b(ws, r, 8, name)
    cl = b(ws, r, 9, formula)
    cl.font = BLD


# ── Save ──────────────────────────────────────────────────────

output = "/home/marccalvet/SlotDemo/docs/total-bet-model-rtp.xlsx"
wb.save(output)
print(f"OK: {output}")
